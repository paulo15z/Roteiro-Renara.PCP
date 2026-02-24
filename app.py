import os, sqlite3, uuid
from datetime import datetime
from flask import Flask, request, jsonify, send_file, render_template, g
import pandas as pd
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

# Em Docker, dados persistentes ficam em /data (volume montado no host)
# Em dev local, ficam na pasta do projeto
DATA_DIR    = os.environ.get('DATA_DIR', os.path.dirname(os.path.abspath(__file__)))
DB_PATH     = os.path.join(DATA_DIR, 'historico.db')
OUTPUTS_DIR = os.path.join(DATA_DIR, 'outputs')
os.makedirs(OUTPUTS_DIR, exist_ok=True)

# ─── BANCO DE DADOS ──────────────────────────────────────────────────────────
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db: db.close()

def init_db():
    with app.app_context():
        db = get_db()
        db.execute('''
            CREATE TABLE IF NOT EXISTS pedidos (
                id          TEXT PRIMARY KEY,
                nome        TEXT,
                data        TEXT,
                total_pecas INTEGER,
                arquivo_out TEXT
            )
        ''')
        db.commit()

# ─── LÓGICA DE ROTEIRO ───────────────────────────────────────────────────────
# Nomes de coluna do template padronizado Dinabox PCP
BORDA_COLS    = ['BORDA_FRENTE', 'BORDA_TRASEIRA', 'BORDA_LE', 'BORDA_LD']
# Fallback para exports antigos (antes do template padronizado)
BORDA_COLS_V1 = ['BORDA_FACE_FRENTE', 'BORDA_FACE_TRASEIRA', 'BORDA_FACE_LE', 'BORDA_FACE_LD']

def calcular_roteiro(row):
    local     = str(row.get('LOCAL', '')).strip()
    # suporta coluna nova (DESCRICAO DA PECA) e antiga (DESCRIÇÃO DA PEÇA)
    descricao = str(row.get('DESCRICAO DA PECA',
                  row.get('DESCRIÇÃO DA PEÇA', ''))).strip().lower()
    duplagem  = str(row.get('DUPLAGEM', '')).strip().lower()
    furo      = str(row.get('FURO', '')).strip()
    # detecta qual conjunto de colunas de borda está presente
    bcols = BORDA_COLS if any(c in row.index for c in BORDA_COLS) else BORDA_COLS_V1
    tem_borda = any(str(row.get(c, '')).strip() not in ('', 'nan') for c in bcols)
    eh_perfil = 'perfil db' in descricao

    rota = ['COR']

    if tem_borda:
        rota.append('XBOR' if eh_perfil else 'BOR')

    if 'duplagem' in duplagem:
        rota.append('DUP')

    if furo:
        rota.append('USI')
        rota.append('FUR')

    if local in ('Caixa', 'Gaveta'):
        rota.append('CAX')
    elif local == 'Porta':
        rota.append('XMAR' if eh_perfil else 'MAR')
    elif local == 'Tamponamento':
        rota.append('MAR')

    rota.append('EXP')
    return ' > '.join(rota)

# ─── GERAÇÃO DO XLSX ─────────────────────────────────────────────────────────
COR_HEADER  = "1F3864"
COR_ALT     = "EEF2F7"
COR_ROT_BG  = "FFF2CC"
COR_ROT_HD  = "BF8F00"

# Colunas do template padronizado (novo)
COLUNAS_EXIB = [
    'ID DO PROJETO', 'NOME DO PROJETO', 'DESCRICAO MODULO',
    'DESCRICAO DA PECA', 'ID DA PECA', 'QUANTIDADE',
    'LARGURA', 'ALTURA', 'ESPESSURA',
    'MATERIAL', 'LOCAL',
    'BORDA_FRENTE', 'BORDA_TRASEIRA', 'BORDA_LE', 'BORDA_LD',
    'FURO', 'DUPLAGEM', 'ROTEIRO'
]
# Fallback para exports antigos
COLUNAS_EXIB_V1 = [
    'ID DO PROJETO', 'NOME DO PROJETO', 'DESCRIÇÃO MÓDULO',
    'DESCRIÇÃO DA PEÇA', 'ID DA PEÇA', 'QUANTIDADE',
    'LARGURA DA PEÇA', 'ALTURA DA PEÇA', 'ESPESSURA',
    'MATERIAL DA PEÇA', 'LOCAL',
    'BORDA_FACE_FRENTE', 'BORDA_FACE_TRASEIRA', 'BORDA_FACE_LE', 'BORDA_FACE_LD',
    'FURO', 'DUPLAGEM', 'ROTEIRO'
]

LARGURAS = {
    # template novo
    'ID DO PROJETO': 14, 'NOME DO PROJETO': 20, 'DESCRICAO MODULO': 22,
    'DESCRICAO DA PECA': 28, 'ID DA PECA': 12, 'QUANTIDADE': 8,
    'LARGURA': 10, 'ALTURA': 10, 'ESPESSURA': 8,
    'MATERIAL': 16, 'LOCAL': 14,
    'BORDA_FRENTE': 14, 'BORDA_TRASEIRA': 14,
    'BORDA_LE': 14, 'BORDA_LD': 14,
    # template antigo (fallback)
    'DESCRIÇÃO MÓDULO': 22, 'DESCRIÇÃO DA PEÇA': 28, 'ID DA PEÇA': 12,
    'LARGURA DA PEÇA': 10, 'ALTURA DA PEÇA': 10, 'MATERIAL DA PEÇA': 16,
    'BORDA_FACE_FRENTE': 14, 'BORDA_FACE_TRASEIRA': 14,
    'BORDA_FACE_LE': 14, 'BORDA_FACE_LD': 14,
    # comum
    'FURO': 7, 'DUPLAGEM': 16, 'ROTEIRO': 42
}

def gerar_xlsx(df):
    # detecta automaticamente template novo ou antigo
    cols = COLUNAS_EXIB if 'DESCRICAO DA PECA' in df.columns else COLUNAS_EXIB_V1

    wb = Workbook()
    ws = wb.active
    ws.title = "Roteiro de Peças"
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cabeçalho
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill  = PatternFill("solid", start_color=COR_ROT_HD if col == 'ROTEIRO' else COR_HEADER)
        c.font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws.row_dimensions[1].height = 35

    # Linhas
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        bg = COR_ALT if ri % 2 == 0 else 'FFFFFF'
        for ci, col in enumerate(cols, 1):
            val = str(row.get(col, '')).strip()
            val = '' if val in ('nan', 'NaN') else val
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col == 'ROTEIRO':
                c.fill = PatternFill("solid", start_color=COR_ROT_BG)
                c.font = Font(name='Arial', size=9, bold=True)
            else:
                c.fill = PatternFill("solid", start_color=bg)
                c.font = Font(name='Arial', size=9)

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = LARGURAS.get(col, 14)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # Aba legenda
    wl = wb.create_sheet("Legenda")
    legenda = [
        ("SETOR", "NOME", "CRITÉRIO"),
        ("COR",  "Corte",               "Toda peça"),
        ("BOR",  "Bordo automático",    "Tem borda, sem 'Perfil db'"),
        ("XBOR", "Bordo manual",        "Tem borda + 'Perfil db' na descrição"),
        ("DUP",  "Duplagem",            "Coluna DUPLAGEM preenchida"),
        ("USI",  "Usinagem",            "Coluna FURO preenchida"),
        ("FUR",  "Furação",             "Coluna FURO preenchida"),
        ("CAX",  "Caixas",              "LOCAL = Caixa ou Gaveta"),
        ("MAR",  "Marcenaria",          "LOCAL = Porta (sem Perfil db) ou Tamponamento"),
        ("XMAR", "Marcenaria especial", "LOCAL = Porta com 'Perfil db'"),
        ("EXP",  "Expedição",           "Toda peça"),
    ]
    wl.column_dimensions['A'].width = 10
    wl.column_dimensions['B'].width = 24
    wl.column_dimensions['C'].width = 42
    for ri, row in enumerate(legenda, 1):
        for ci, val in enumerate(row, 1):
            c = wl.cell(row=ri, column=ci, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='left', vertical='center')
            if ri == 1:
                c.fill = PatternFill("solid", start_color=COR_HEADER)
                c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            else:
                c.fill = PatternFill("solid", start_color=COR_ALT if ri % 2 == 0 else 'FFFFFF')
                c.font = Font(name='Arial', size=10)
        wl.row_dimensions[ri].height = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─── PROCESSAMENTO DO ARQUIVO ────────────────────────────────────────────────
def processar_arquivo(file):
    """
    Lê CSV exportado pelo Dinabox (separador ;, encoding latin-1).
    Suporta arquivos com blocos [CABECALHO]/[LISTA]/[RODAPE].
    Remove o ponto-e-vírgula extra que o Dinabox coloca no final de cada linha.
    Também aceita XLSX como fallback.
    """
    import tempfile, os as _os
    nome = file.filename
    ext  = nome.rsplit('.', 1)[-1].lower()

    if ext == 'csv':
        raw = file.read()
        text = None
        for enc in ('utf-8-sig', 'latin-1', 'cp1252', 'utf-8'):
            try:
                text = raw.decode(enc)
                break
            except Exception:
                continue
        if text is None:
            raise ValueError("Não foi possível decodificar o arquivo CSV.")

        linhas = text.splitlines()
        cabecalho = []
        dados = []
        em_bloco = None

        for linha in linhas:
            s = linha.strip()
            if s == '[CABECALHO]':   em_bloco = 'cab'; continue
            if s == '[/CABECALHO]':  em_bloco = None;  continue
            if s == '[LISTA]':       em_bloco = 'lst'; continue
            if s == '[/LISTA]':      em_bloco = None;  continue
            if s.startswith('['):    em_bloco = None;  continue
            if not s:                continue
            if em_bloco == 'cab':    cabecalho.append(linha.rstrip(';'))
            elif em_bloco == 'lst':  dados.append(linha.rstrip(';'))

        if not cabecalho:
            # Arquivo sem blocos de controle — usa as linhas diretamente
            todas = [l.rstrip(';') for l in linhas
                     if l.strip() and l.strip() not in ('RODAPÉ',)]
        else:
            todas = cabecalho + dados

        df = pd.read_csv(StringIO('\n'.join(todas)), sep=';', dtype=str,
                         skipinitialspace=True)

    elif ext == 'xlsx':
        df = pd.read_excel(file, dtype=str, engine='openpyxl')

    elif ext == 'xls':
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xls')
        try:
            file.save(tmp.name)
            tmp.close()
            df = pd.read_excel(tmp.name, dtype=str, engine='xlrd')
        except Exception as e:
            raise ValueError(
                f"Erro ao ler .xls: {e}\n"
                "Dica: abra no Excel e salve como .xlsx, depois tente novamente."
            )
        finally:
            _os.unlink(tmp.name)

    else:
        raise ValueError("Formato não suportado. Use CSV, XLS ou XLSX.")

    df = df.fillna('')

    # Normaliza nomes de colunas (remove espaços extras)
    df.columns = [c.strip() for c in df.columns]
    # Remove coluna vazia gerada pelo ; extra (se ainda sobrar)
    df = df[[c for c in df.columns if c]]

    # Remove linha de rodapé residual
    if 'NOME DO CLIENTE' in df.columns:
        df = df[~df['NOME DO CLIENTE'].str.strip().isin(['RODAPÉ', ''])]

    # Verifica colunas mínimas
    obrigatorias = ['LOCAL', 'FURO', 'DUPLAGEM', 'DESCRIÇÃO DA PEÇA']
    faltando = [c for c in obrigatorias if c not in df.columns]
    if faltando:
        raise ValueError(
            f"Colunas não encontradas: {', '.join(faltando)}\n"
            "Use o template_dinabox_pcp.txt para configurar a exportação."
        )

    df['ROTEIRO'] = df.apply(calcular_roteiro, axis=1)
    return df


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/processar', methods=['POST'])
def processar():
    if 'arquivo' not in request.files:
        return jsonify({'erro': 'Nenhum arquivo enviado.'}), 400

    file = request.files['arquivo']
    if not file.filename:
        return jsonify({'erro': 'Arquivo vazio.'}), 400

    try:
        df = processar_arquivo(file)
    except ValueError as e:
        return jsonify({'erro': str(e)}), 400
    except Exception as e:
        return jsonify({'erro': f'Erro ao processar arquivo: {str(e)}'}), 500

    # Monta prévia para o frontend
    previa = df[['DESCRIÇÃO DA PEÇA', 'LOCAL', 'ROTEIRO']].head(50).to_dict(orient='records')

    # Resumo por roteiro
    resumo = df['ROTEIRO'].value_counts().reset_index()
    resumo.columns = ['roteiro', 'qtd']
    resumo_list = resumo.to_dict(orient='records')

    # Salva XLSX
    pid     = str(uuid.uuid4())[:8]
    nome_arq = f"{pid}_{file.filename.rsplit('.', 1)[0]}.xlsx"
    caminho = os.path.join(OUTPUTS_DIR, nome_arq)
    xlsx_buf = gerar_xlsx(df)
    with open(caminho, 'wb') as f:
        f.write(xlsx_buf.read())

    # Salva no histórico
    db = get_db()
    db.execute(
        'INSERT INTO pedidos VALUES (?,?,?,?,?)',
        (pid, file.filename, datetime.now().strftime('%d/%m/%Y %H:%M'), len(df), nome_arq)
    )
    db.commit()

    return jsonify({
        'pid':    pid,
        'total':  len(df),
        'previa': previa,
        'resumo': resumo_list,
    })

@app.route('/download/<pid>')
def download(pid):
    db  = get_db()
    row = db.execute('SELECT * FROM pedidos WHERE id=?', (pid,)).fetchone()
    if not row:
        return 'Não encontrado', 404
    caminho = os.path.join(OUTPUTS_DIR, row['arquivo_out'])
    return send_file(caminho, as_attachment=True,
                     download_name=f"ROTEIRO_{row['nome']}.xlsx")

@app.route('/historico')
def historico():
    db   = get_db()
    rows = db.execute('SELECT * FROM pedidos ORDER BY data DESC').fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/historico/<pid>', methods=['DELETE'])
def deletar(pid):
    db  = get_db()
    row = db.execute('SELECT arquivo_out FROM pedidos WHERE id=?', (pid,)).fetchone()
    if row:
        try:
            os.remove(os.path.join(OUTPUTS_DIR, row['arquivo_out']))
        except FileNotFoundError:
            pass
        db.execute('DELETE FROM pedidos WHERE id=?', (pid,))
        db.commit()
    return jsonify({'ok': True})

if __name__ == '__main__':
    with app.app_context():
        init_db()
    debug = os.environ.get('FLASK_DEBUG', '0') == '1'
    host  = '127.0.0.1' if debug else '0.0.0.0'
    app.run(host=host, port=5000, debug=debug)